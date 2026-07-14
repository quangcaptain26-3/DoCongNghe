#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AGV 日報整合工具
  - 異常率：讀取 abnormalAnalyse.py 產出的 XLSX（OutputLog/Log{date}/Log{date}_day.xlsx / _night.xlsx）
  - 稼動率：讀取任務原始 CSV，以 complete_time 判斷日/夜班歸屬
  - 輸出：{date}_report.csv（日班 + 夜班，各含稼動率與異常分析）

設定檔（與本程式同目錄）：report_setting.ini
執行方式：python reportAnalyse.py
"""

import configparser
import csv as csv_mod
import re
import sys
from datetime import datetime, timedelta
from collections import defaultdict
from pathlib import Path

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False


# ── Helpers ───────────────────────────────────────────────────────────────────

def get_script_dir() -> Path:
    if getattr(sys, 'frozen', False):
        return Path(sys.executable).parent
    return Path(__file__).parent


# ── Settings ──────────────────────────────────────────────────────────────────

def load_settings(script_dir: Path):
    ini_path = script_dir / "report_setting.ini"
    if not ini_path.exists():
        print(f"錯誤: 找不到設定檔 {ini_path}")
        return None

    config = configparser.ConfigParser()
    config.read(ini_path, encoding='utf-8-sig')

    try:
        abnormal_path = config.get('paths', 'abnormalOutputPath').strip("'\"")
        task_path     = config.get('paths', 'taskCsvPath').strip("'\"")
        output_path   = config.get('paths', 'outputPath').strip("'\"")

        day_hs  = config.getint('shift', 'day_hr_start')
        day_ms  = config.getint('shift', 'day_min_start')
        day_he  = config.getint('shift', 'day_hr_end')
        day_me  = config.getint('shift', 'day_min_end')
        night_hs = config.getint('shift', 'night_hr_start')
        night_ms = config.getint('shift', 'night_min_start')
        night_he = config.getint('shift', 'night_hr_end')
        night_me = config.getint('shift', 'night_min_end')
    except (configparser.NoSectionError, configparser.NoOptionError) as e:
        print(f"錯誤: 設定檔格式錯誤 - {e}")
        return None

    return {
        'abnormal_path': abnormal_path,
        'task_path':     task_path,
        'output_path':   output_path,
        'shift_config': {
            'day':   (day_hs, day_ms, day_he, day_me),
            'night': (night_hs, night_ms, night_he, night_me),
        },
    }


def calculate_shift_ranges(shift_config):
    """
    計算昨日日班與夜班的精確時間範圍（含秒數，確保 12H 精確）。
    Day:   yesterday 08:00:00 ~ 19:59:59
    Night: yesterday 20:00:00 ~ today   07:59:59
    """
    yesterday = datetime.now() - timedelta(days=1)
    today     = datetime.now()

    d_hs, d_ms, d_he, d_me = shift_config['day']
    n_hs, n_ms, n_he, n_me = shift_config['night']

    day_start   = yesterday.replace(hour=d_hs, minute=d_ms, second=0,  microsecond=0)
    day_end     = yesterday.replace(hour=d_he, minute=d_me, second=59, microsecond=999999)
    night_start = yesterday.replace(hour=n_hs, minute=n_ms, second=0,  microsecond=0)
    night_end   = today.replace(    hour=n_he, minute=n_me, second=59, microsecond=999999)

    return day_start, day_end, night_start, night_end


# ── 異常分析：讀取 XLSX ────────────────────────────────────────────────────────

def find_abnormal_xlsx(abnormal_path: str, date_prefix: str):
    """在 abnormal_path 底下遞迴搜尋 _day.xlsx 與 _night.xlsx（不限資料夾層數）"""
    base = Path(abnormal_path)
    if not base.exists():
        print(f"  找不到異常根目錄: {base}")
        return None, None

    day_name   = f"Log{date_prefix}_day.xlsx"
    night_name = f"Log{date_prefix}_night.xlsx"

    day_matches   = list(base.rglob(day_name))
    night_matches = list(base.rglob(night_name))

    day_xlsx   = day_matches[0]   if day_matches   else None
    night_xlsx = night_matches[0] if night_matches else None

    if not day_xlsx:
        print(f"  找不到日班 XLSX: {day_name}")
    else:
        print(f"  找到日班 XLSX: {day_xlsx}")
    if not night_xlsx:
        print(f"  找不到夜班 XLSX: {night_name}")
    else:
        print(f"  找到夜班 XLSX: {night_xlsx}")

    return day_xlsx, night_xlsx


def read_abnormal_xlsx(xlsx_path: Path):
    """
    解析 abnormalAnalyse.py 輸出的 XLSX，提取車輛異常記錄。

    XLSX 資料結構（固定格式）：
      Row 8+  B=車號  C=異常次數  D=點位+時間  E=停留分鐘
      D 欄格式：「點位{id} / {HH:MM:SS}」或「-」
      無異常的車輛：C=0  D='-'  E='-'

    Returns: dict  { car_id: [ {'point_id', 'arrival_time', 'stay_min'}, ... ] }
      car_id 存在但 list 為空 → 該車無異常記錄
    """
    if not HAS_OPENPYXL:
        print("  錯誤: 需要 openpyxl。請執行 'pip install openpyxl'")
        return {}

    result = {}
    current_car = None

    try:
        wb = load_workbook(xlsx_path, data_only=True)
        ws = wb.active

        for row in ws.iter_rows(min_row=8, values_only=True):
            b_val = row[1]   # Column B：車號
            d_val = row[3]   # Column D：點位/時間
            e_val = row[4]   # Column E：停留分鐘

            # 完全空行（資料結尾後的空行）→ 略過
            if b_val is None and d_val is None:
                continue

            # B 欄有值 → 可能是新車輛行 或 備註行
            if b_val is not None and str(b_val).strip():
                b_str = str(b_val).strip()
                # 車號皆為純數字；備註行（如「異常閾值」）不是純數字
                if b_str.isdigit():
                    if b_str == '0':
                        current_car = None   # 排除無效車號 0
                        continue
                    current_car = b_str
                    if current_car not in result:
                        result[current_car] = []
                else:
                    continue   # 備註行，略過

            if current_car is None:
                continue

            # 解析 D 欄：「點位532 / 08:32:15」或「-」
            d_str = str(d_val).strip() if d_val is not None else ''
            if d_str and d_str != '-' and '點位' in d_str and '/' in d_str:
                parts = d_str.split('/')
                point_id     = parts[0].replace('點位', '').strip()
                arrival_time = parts[1].strip()
                try:
                    stay_min = round(float(e_val), 2)
                except (TypeError, ValueError):
                    stay_min = 0.0
                result[current_car].append({
                    'point_id':    point_id,
                    'arrival_time': arrival_time,
                    'stay_min':    stay_min,
                })

    except Exception as e:
        print(f"  警告: 讀取 XLSX {xlsx_path.name} 失敗: {e}")

    return result


# ── 稼動率：讀取任務 CSV ───────────────────────────────────────────────────────

def load_task_df(task_path: str, yesterday_prefix: str, today_prefix: str):
    """
    載入昨日與今日的任務 CSV（排除 _taskTime 結尾），合併成一個 DataFrame。
    於 task_path 底下遞迴搜尋 {date}.csv（不限資料夾層數）。
    自動偵測 comma/tab 分隔符及編碼。
    """
    if not HAS_PANDAS:
        return None

    task_folder = Path(task_path)
    dfs = []

    for prefix in [yesterday_prefix, today_prefix]:
        matches = [p for p in task_folder.rglob(f"{prefix}.csv")
                   if not p.stem.endswith('_taskTime')]
        if not matches:
            print(f"    找不到: {prefix}.csv，略過。")
            continue
        csv_file = matches[0]

        rows = None
        for enc in ('utf-8-sig', 'utf-8', 'cp950', 'gbk'):
            try:
                with open(csv_file, 'r', encoding=enc, newline='') as f:
                    sample = f.read(4096)
                    try:
                        dialect   = csv_mod.Sniffer().sniff(sample, delimiters=',\t')
                        delimiter = dialect.delimiter
                    except csv_mod.Error:
                        delimiter = ','
                    f.seek(0)
                    reader = csv_mod.DictReader(f, delimiter=delimiter, restkey='__overflow__')
                    rows = [{k: v for k, v in r.items() if k != '__overflow__'} for r in reader]
                break
            except UnicodeDecodeError:
                continue

        if rows:
            df = pd.DataFrame(rows)
            df.columns = [c.strip() if isinstance(c, str) else c for c in df.columns]
            dfs.append(df)
            print(f"    載入: {csv_file}（{len(rows)} 筆）")

    return pd.concat(dfs, ignore_index=True) if dfs else None


def parse_dt_flexible(series):
    """彈性解析時間字串：支援帶秒/不帶秒、連字號/斜線等格式"""
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S',
                '%Y-%m-%d %H:%M',    '%Y/%m/%d %H:%M'):
        try:
            return pd.to_datetime(series, format=fmt, errors='raise')
        except (ValueError, TypeError):
            continue
    return pd.to_datetime(series, errors='coerce')   # 最後讓 pandas 自行推斷


def analyse_utilization_by_shift(df, day_start, day_end, night_start, night_end):
    """
    依班別計算稼動率。
    使用 complete_time 判斷任務歸屬班別；只計 final_state == 'completed'。
    稼動率 = 各車 completed duration 合計 ÷ (班別時長(秒) × 台數) × 100

    Returns: {
        'day':   {'car_rows': list, 'utilization': float, 'car_count': int, 'shift_h': float} | None,
        'night': { ... }
    }
    """
    required = {'carId', 'duration_sec', 'final_state', 'complete_time'}
    missing  = required - set(df.columns)
    if missing:
        print(f"  警告: task CSV 缺少欄位 {missing}，稼動率分析跳過。")
        return None

    df = df[df['final_state'].str.strip() == 'completed'].copy()
    # 排除無效車號（0、空值）
    df['carId'] = df['carId'].astype(str).str.strip()
    df = df[~df['carId'].str.lower().isin(['0', '', 'nan', 'none'])]
    df['duration_sec'] = pd.to_numeric(df['duration_sec'], errors='coerce')
    df = df.dropna(subset=['duration_sec'])

    if df.empty:
        print("  警告: 無 completed 記錄。")
        return None

    df['complete_dt'] = parse_dt_flexible(df['complete_time'].astype(str).str.strip())
    df = df.dropna(subset=['complete_dt'])

    results = {}
    for shift_key, (s_start, s_end) in [('day', (day_start, day_end)),
                                         ('night', (night_start, night_end))]:
        shift_sec = (s_end - s_start).total_seconds()
        shift_h   = round(shift_sec / 3600, 1)

        sdf = df[(df['complete_dt'] >= s_start) & (df['complete_dt'] <= s_end)]

        if sdf.empty:
            results[shift_key] = None
            continue

        grouped = (sdf.groupby('carId')['duration_sec']
                   .agg(task_count='count', total_sec='sum')
                   .reset_index()
                   .sort_values('carId'))

        car_count   = len(grouped)
        utilization = grouped['total_sec'].sum() / (shift_sec * car_count) * 100

        results[shift_key] = {
            'car_rows': [{
                'carId':       str(r['carId']),
                'task_count':  int(r['task_count']),
                'total_sec':   round(float(r['total_sec']), 1),
                'total_hours': round(float(r['total_sec']) / 3600, 4),
            } for _, r in grouped.iterrows()],
            'utilization': round(utilization, 2),
            'car_count':   car_count,
            'shift_h':     shift_h,
        }

    return results


# ── 輸出整合 CSV ───────────────────────────────────────────────────────────────

def export_report(date_str, day_start, day_end, night_start, night_end,
                  abnormal_day, abnormal_night, util_results, output_folder: Path):
    """
    輸出 {date}_report.csv：
      各班別依序：稼動率表 → 異常分析表
    """
    out_file = output_folder / f"{date_str}_report.csv"

    shift_meta = {
        'day':   (day_start, day_end,   '日班', abnormal_day),
        'night': (night_start, night_end, '夜班', abnormal_night),
    }

    with open(out_file, 'w', encoding='utf-8-sig', newline='') as f:
        w = csv_mod.writer(f)

        w.writerow(['AGV 日報整合報告'])
        w.writerow(['日期', date_str])
        w.writerow([])

        for shift_key in ['day', 'night']:
            s_start, s_end, label, abn_data = shift_meta[shift_key]

            w.writerow([f'=== {label}  {s_start.strftime("%H:%M")} ~ {s_end.strftime("%H:%M")} ==='])
            w.writerow([])

            # ── 稼動率 ────────────────────────────────────────────────
            w.writerow(['【稼動率】'])
            util = (util_results or {}).get(shift_key)

            if util:
                w.writerow(['車號', '任務次數', '總時間(秒)', '總時間(小時)'])
                for r in util['car_rows']:
                    w.writerow([r['carId'], r['task_count'], r['total_sec'], r['total_hours']])
                w.writerow([])
                w.writerow(['總稼動率', f"{util['utilization']}%",
                            f"（基準：{util['shift_h']}H × {util['car_count']} 台）"])
            else:
                w.writerow(['（無資料）'])

            w.writerow([])

            # ── 異常分析 ───────────────────────────────────────────────
            w.writerow(['【異常分析】'])

            if abn_data:
                w.writerow(['車號', '異常次數', '異常點位', '抵達時間', '停留時間(分鐘)'])
                for car_id in sorted(abn_data.keys()):
                    recs = abn_data[car_id]
                    if recs:
                        for j, rec in enumerate(recs):
                            w.writerow([
                                car_id if j == 0 else '',
                                len(recs) if j == 0 else '',
                                f"點位{rec['point_id']}",
                                rec['arrival_time'],
                                rec['stay_min'],
                            ])
                    else:
                        w.writerow([car_id, 0, '-', '-', '-'])
            else:
                w.writerow(['（無資料）'])

            w.writerow([])
            w.writerow([])

    print(f"  整合報告已輸出至: {out_file}")
    return out_file


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("AGV 日報整合工具（稼動率 + 異常分析）")
    print("=" * 60)

    if not HAS_OPENPYXL:
        print("錯誤: 需要 openpyxl。請執行 'pip install openpyxl'")
        return 1
    if not HAS_PANDAS:
        print("警告: 未安裝 pandas，稼動率分析將跳過。請執行 'pip install pandas'")

    script_dir = get_script_dir()

    # 1. 讀取設定
    print("\n[1] 讀取設定檔...")
    cfg = load_settings(script_dir)
    if not cfg:
        return 1

    abnormal_path = cfg['abnormal_path']
    task_path     = cfg['task_path']
    output_path   = cfg['output_path']
    shift_config  = cfg['shift_config']
    print(f"  異常 XLSX 根目錄: {abnormal_path}")
    print(f"  任務 CSV 目錄:    {task_path}")
    print(f"  輸出目錄:         {output_path}")

    # 2. 計算時間範圍
    day_start, day_end, night_start, night_end = calculate_shift_ranges(shift_config)
    yesterday_prefix = (datetime.now() - timedelta(days=1)).strftime("%Y%m%d")
    today_prefix     = datetime.now().strftime("%Y%m%d")

    print(f"\n[2] 分析日期: {yesterday_prefix}")
    print(f"  日班: {day_start.strftime('%H:%M')} ~ {day_end.strftime('%H:%M')}")
    print(f"  夜班: {night_start.strftime('%H:%M')} ~ {night_end.strftime('%H:%M')} "
          f"（含跨日至 {night_end.strftime('%Y-%m-%d')}）")

    # 3. 讀取異常 XLSX
    print(f"\n[3] 讀取異常分析 XLSX...")
    day_xlsx, night_xlsx = find_abnormal_xlsx(abnormal_path, yesterday_prefix)

    abnormal_day   = read_abnormal_xlsx(day_xlsx)   if day_xlsx   else {}
    abnormal_night = read_abnormal_xlsx(night_xlsx) if night_xlsx else {}

    day_abn_total   = sum(len(v) for v in abnormal_day.values())
    night_abn_total = sum(len(v) for v in abnormal_night.values())
    print(f"  日班：{len(abnormal_day)} 台車，{day_abn_total} 筆異常")
    print(f"  夜班：{len(abnormal_night)} 台車，{night_abn_total} 筆異常")

    # 4. 讀取任務 CSV，依班別計算稼動率
    print(f"\n[4] 讀取任務 CSV（{yesterday_prefix} + {today_prefix}）...")
    util_results = None
    if HAS_PANDAS:
        task_df = load_task_df(task_path, yesterday_prefix, today_prefix)
        if task_df is not None:
            print("  計算班別稼動率...")
            util_results = analyse_utilization_by_shift(
                task_df, day_start, day_end, night_start, night_end)
            if util_results:
                for sk, sl in [('day', '日班'), ('night', '夜班')]:
                    u = util_results.get(sk)
                    if u:
                        print(f"  {sl} 稼動率: {u['utilization']}%（{u['car_count']} 台車）")
                    else:
                        print(f"  {sl}: 無 completed 記錄")
        else:
            print("  找不到任務 CSV，稼動率跳過。")

    # 5. 輸出整合報告
    print(f"\n[5] 輸出整合報告...")
    output_folder = Path(output_path) / f"Log{yesterday_prefix}"
    output_folder.mkdir(parents=True, exist_ok=True)

    export_report(
        date_str       = yesterday_prefix,
        day_start      = day_start,
        day_end        = day_end,
        night_start    = night_start,
        night_end      = night_end,
        abnormal_day   = abnormal_day,
        abnormal_night = abnormal_night,
        util_results   = util_results,
        output_folder  = output_folder,
    )

    print("\n" + "=" * 60)
    print("完成!")
    print("=" * 60)
    return 0


if __name__ == '__main__':
    sys.exit(main())
