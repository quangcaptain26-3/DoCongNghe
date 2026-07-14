#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AGV 異常停留統計工具 (日班+夜班)
統計每台 AGV 的異常停留點位，輸出日班/夜班各一份 Excel

執行方式：
  python logAnalyse_by_daynight.py
  自動讀取 abnormal_setting.ini 與 point_settings.json，處理前一天的 log 檔案
  產出日班/夜班各一份 Excel (僅 AGV 異常時間統計)
"""

import configparser
import json
import re
import shutil
import sys
from datetime import datetime, timedelta
from collections import defaultdict
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("警告: 未安裝 openpyxl，請執行 'pip install openpyxl' 以啟用 Excel 輸出。")


# 時間戳記格式
TIMESTAMP_PATTERN = r'(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2}\.\d{3})'

# 充電相關（僅保留狀態追蹤所需，不計算時間）
CHARGE_START_PATTERN   = re.compile(r'(\d{4})标记自动充电(\d+)')
CHARGE_END_PATTERN     = re.compile(r'(\d{4})充电结束')
CHARGE_RELEASE_PATTERN = re.compile(r'(\d{4})缓存释放，充电桩释放\d+')

# 任務相關（僅保留狀態追蹤所需，不計算時間）
TASK_ASSIGN_PATTERN  = re.compile(r'任务([\w-]+)分配给(\d{4})号车')
TASK_FINISH_PATTERN  = re.compile(r'(\d{4})车辆状态修改为NONE，FINISH')

# 點位相關
POINT_ARRIVAL_PATTERN = re.compile(r'(\d{4})目标点[：:](\d+)')


def parse_timestamp(line):
    """從 log 行解析時間戳記"""
    match = re.match(TIMESTAMP_PATTERN, line)
    if match:
        ts_str = re.sub(r'\s+', ' ', match.group(1).strip())
        try:
            return datetime.strptime(ts_str, "%Y-%m-%d %H:%M:%S.%f")
        except ValueError:
            return None
    return None


def determine_shift(timestamp, day_range, night_range):
    """判斷時間戳屬於哪個班別；Returns: 'day', 'night', or None"""
    day_start, day_end = day_range
    night_start, night_end = night_range
    if day_start <= timestamp <= day_end:
        return 'day'
    if night_start <= timestamp <= night_end:
        return 'night'
    return None


def load_point_settings(script_dir):
    """
    從 point_settings.json 載入特定點位設定。

    JSON 結構:
      {
        "excluded_points": {
          "Home":     ["232", ...],   ← 完全排除異常統計（停留再久不算異常）
          "Charging": ["239"],
          "standby":  ["501"]         ← 分類名稱可自由新增，程式自動納入
        },
        "elevator": {
          "threshold_min": 3,         ← 電梯點位專用閾值（使用者可自行調整）
          "points": ["221", "1294", "223", "1767"]
        }
      }

    Returns dict:
      {
        'excluded_points':       set of str,
        'elevator_points':       set of str,
        'elevator_threshold_sec': float,
      }
    """
    defaults = {
        'excluded_points':        set(),
        'elevator_points':        set(),
        'elevator_threshold_sec': 180.0,
        'abnormal_threshold_min': 12,
    }
    path = script_dir / "point_settings.json"
    if not path.exists():
        print("  提示: 找不到 point_settings.json，所有點位使用預設設定。")
        return defaults
    try:
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        excluded = set()
        for category, points in data.get('excluded_points', {}).items():
            if category.startswith('_'):
                continue
            for p in points:
                excluded.add(str(p))

        elevator_data          = data.get('elevator', {})
        elevator_threshold_min = elevator_data.get('threshold_min', 3)
        elevator_pts           = set(str(p) for p in elevator_data.get('points', []))

        abnormal_threshold_min = data.get('threshold_min', 12)
        print(f"  一般異常閾值: {abnormal_threshold_min} 分鐘；排除點位: {len(excluded)} 個；電梯點位: {len(elevator_pts)} 個 (閾值 {elevator_threshold_min} 分鐘)")
        return {
            'excluded_points':         excluded,
            'elevator_points':         elevator_pts,
            'elevator_threshold_sec':  elevator_threshold_min * 60,
            'abnormal_threshold_min':  abnormal_threshold_min,
        }
    except Exception as e:
        print(f"  警告: 讀取 point_settings.json 失敗: {e}")
        return defaults


def parse_log_files_by_shift(log_files, day_range, night_range, abnormal_threshold_sec,
                             excluded_points, elevator_points, elevator_threshold_sec):
    """
    解析所有 log 檔案，僅統計異常停留事件，依發生時間歸屬班別。

    Returns:
        {
            'day':   (abnormal_records_list, seen_cars_set),
            'night': (abnormal_records_list, seen_cars_set),
        }
    """
    day_start, day_end = day_range
    night_start, night_end = night_range

    car_states = defaultdict(lambda: {
        'is_charging': False,
        'is_task_active': False,
        'last_point': None,
        'last_point_time': None,
        'last_point_had_task': False,
    })

    shift_abnormal  = {'day': [], 'night': []}
    shift_seen_cars = {'day': set(), 'night': set()}

    def check_abnormal(car_id, new_point_time, shift):
        """檢查上一個點位是否異常停留（排除點位不計入；電梯點位用獨立閾值）"""
        state = car_states[car_id]
        if state['last_point'] is None or state['last_point_time'] is None:
            return
        point_str = str(state['last_point'])
        if point_str in excluded_points:
            return
        threshold = elevator_threshold_sec if point_str in elevator_points else abnormal_threshold_sec
        stay_duration = (new_point_time - state['last_point_time']).total_seconds()
        if (stay_duration > threshold
                and not state['is_charging']
                and state['last_point_had_task']):
            if shift and shift in shift_abnormal:
                shift_abnormal[shift].append({
                    'car_id': car_id,
                    'point_id': state['last_point'],
                    'arrival_time': state['last_point_time'],
                    'stay_duration': stay_duration,
                })

    for log_file in log_files:
        try:
            with open(log_file, 'r', encoding='utf-8') as f:
                for line in f:
                    line = line.strip()
                    if not line:
                        continue

                    timestamp = parse_timestamp(line)
                    if not timestamp:
                        continue

                    if timestamp < day_start or timestamp > night_end:
                        continue

                    # 1. 開始走向充電樁
                    match = CHARGE_START_PATTERN.search(line)
                    if match:
                        car_id = match.group(1)
                        car_states[car_id]['is_charging'] = True
                        continue

                    # 2. 充電結束
                    match = CHARGE_END_PATTERN.search(line)
                    if match:
                        car_states[match.group(1)]['is_charging'] = False
                        continue

                    # 3. 充電樁緩存釋放（視為充電結束）
                    match = CHARGE_RELEASE_PATTERN.search(line)
                    if match:
                        car_states[match.group(1)]['is_charging'] = False
                        continue

                    # 4. 任務分配
                    match = TASK_ASSIGN_PATTERN.search(line)
                    if match:
                        car_id = match.group(2)
                        if car_states[car_id]['is_charging']:
                            car_states[car_id]['is_charging'] = False
                        car_states[car_id]['is_task_active'] = True
                        continue

                    # 5. 整體任務最終完成
                    match = TASK_FINISH_PATTERN.search(line)
                    if match:
                        car_id = match.group(1)
                        car_states[car_id]['is_task_active'] = False
                        car_states[car_id]['last_point_had_task'] = False
                        continue

                    # 6. 到達點位（異常統計）
                    match = POINT_ARRIVAL_PATTERN.search(line)
                    if match:
                        car_id  = match.group(1)
                        point_id = match.group(2)
                        shift = determine_shift(timestamp, day_range, night_range)
                        if shift:
                            shift_seen_cars[shift].add(car_id)
                        check_abnormal(car_id, timestamp, shift)
                        car_states[car_id]['last_point_had_task'] = car_states[car_id]['is_task_active']
                        car_states[car_id]['last_point'] = point_id
                        car_states[car_id]['last_point_time'] = timestamp
                        continue

        except Exception as e:
            print(f"警告: 讀取檔案 {log_file} 時發生錯誤: {e}")

    return {
        'day':   (shift_abnormal['day'],   shift_seen_cars['day']),
        'night': (shift_abnormal['night'], shift_seen_cars['night']),
    }


def print_report(abnormal_records, log_file_count, folder_path, shift_label='', abnormal_threshold_min=12):
    """輸出終端報告（僅異常統計）"""
    print("\n" + "=" * 60)
    title = "AGV 異常停留統計報告"
    if shift_label:
        title += f" ({shift_label})"
    print(title)
    print("=" * 60)
    print(f"統計路徑: {folder_path}")
    print(f"Log 檔案數量: {log_file_count}")
    print(f"異常閾值: {abnormal_threshold_min} 分鐘")
    print("-" * 60)
    print(f"異常記錄數: {len(abnormal_records)}")
    if abnormal_records:
        car_abnormal_count = defaultdict(int)
        for record in abnormal_records:
            car_abnormal_count[record['car_id']] += 1
        for car_id in sorted(car_abnormal_count.keys()):
            print(f"  車號 {car_id}: {car_abnormal_count[car_id]} 次異常")
    print("=" * 60 + "\n")


def export_excel_abnormal(abnormal_records, seen_cars, output_path, shift_label='',
                          abnormal_threshold_min=12, date_prefix='', shift_suffix=''):
    """
    輸出 Excel 檔案 — 僅 AGV 車異常時間統計
    欄位配置: B=車號, C=異常停留次數, D=異常點位, E=異常滯留時間
    """
    if not HAS_OPENPYXL:
        print("錯誤: 需要安裝 openpyxl 才能輸出 Excel。請執行 'pip install openpyxl'")
        return None

    excel_file = Path(output_path) / f"Log{date_prefix}_{shift_suffix}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "daily"

    center    = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_wrap = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    bold      = Font(bold=True, size=11)
    normal    = Font(size=10)
    header_fill = PatternFill('solid', fgColor='D9E1F2')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )

    def style_cell(cell, font=normal, alignment=center, border=thin_border, fill=None):
        cell.font = font
        cell.alignment = alignment
        cell.border = border
        if fill:
            cell.fill = fill

    # Row 1: 標題
    ws.merge_cells('B1:E1')
    c = ws['B1']
    c.value = f"AGV車異常時間統計 ({shift_label})"
    style_cell(c, font=Font(bold=True, size=14))

    # Row 3: 總異常率
    ws['B3'] = '總異常率(O)'
    style_cell(ws['B3'], font=bold)

    ws.merge_cells('D3:E3')
    ws['D3'] = '*所有車的(E)加總/(24H*總車數)'
    style_cell(ws['D3'], font=Font(size=9, italic=True))

    # Row 5: 子標題
    ws['C5'] = '異常停留次數'
    ws['D5'] = '異常點位'
    ws['E5'] = '異常滯留時間'
    for col in ['C', 'D', 'E']:
        style_cell(ws[f'{col}5'], font=bold, fill=header_fill)

    # Row 6: 代號
    ws['C6'] = 'C'
    ws['D6'] = 'D'
    ws['E6'] = 'E'
    for col in ['C', 'D', 'E']:
        style_cell(ws[f'{col}6'], font=bold, fill=header_fill)

    # Row 7: 欄位說明
    ws['B7'] = '車號\\定義'
    ws['C7'] = f'任何點位停留超過{abnormal_threshold_min}min(充電除外)'
    ws['D7'] = '異常點位註記(一個點/一個時間/一筆)'
    ws['E7'] = '前述異常點滯留時間(分鐘)'
    for col in ['B', 'C', 'D', 'E']:
        style_cell(ws[f'{col}7'], font=bold, fill=header_fill)

    # Row 8+: 資料列（按車號分組）
    abnormal_by_car = defaultdict(list)
    for record in abnormal_records:
        abnormal_by_car[record['car_id']].append(record)

    all_cars = sorted(seen_cars | set(abnormal_by_car.keys()))
    data_start_row = 8
    abnormal_row = data_start_row

    for car_id in all_cars:
        records = sorted(abnormal_by_car.get(car_id, []), key=lambda x: x['arrival_time'])
        if records:
            for j, rec in enumerate(records):
                if j == 0:
                    ws.cell(row=abnormal_row, column=2, value=car_id)
                    ws.cell(row=abnormal_row, column=3, value=len(records))
                ws.cell(row=abnormal_row, column=4,
                        value=f"點位{rec['point_id']} / {rec['arrival_time'].strftime('%H:%M:%S')}")
                ws.cell(row=abnormal_row, column=5, value=round(rec['stay_duration'] / 60, 2))
                for col in range(2, 6):
                    style_cell(ws.cell(row=abnormal_row, column=col))
                abnormal_row += 1
        else:
            ws.cell(row=abnormal_row, column=2, value=car_id)
            ws.cell(row=abnormal_row, column=3, value=0)
            ws.cell(row=abnormal_row, column=4, value='-')
            ws.cell(row=abnormal_row, column=5, value='-')
            for col in range(2, 6):
                style_cell(ws.cell(row=abnormal_row, column=col))
            abnormal_row += 1

    # 填入總異常率
    if all_cars:
        total_abnormal_hours = sum(r['stay_duration'] / 3600 for r in abnormal_records)
        total_abnormal_rate  = total_abnormal_hours / (24 * len(all_cars)) * 100
        ws['C3'] = f"{total_abnormal_rate:.2f}%"
        style_cell(ws['C3'], font=Font(bold=True, size=11, color='FF0000'))

    # 備註區
    note_row = abnormal_row + 2
    ws.cell(row=note_row, column=2, value=f'異常閾值: {abnormal_threshold_min} 分鐘 (充電除外)')
    style_cell(ws.cell(row=note_row, column=2),
               font=Font(size=9, italic=True, color='666666'),
               alignment=left_wrap, border=Border())

    # 欄寬調整
    col_widths = {'A': 2, 'B': 12, 'C': 18, 'D': 30, 'E': 20}
    for letter, width in col_widths.items():
        ws.column_dimensions[letter].width = width

    wb.save(excel_file)
    print(f"Excel 報告已輸出至: {excel_file}")
    return excel_file


def load_settings():
    """讀取 abnormal_setting.ini 設定檔"""
    if getattr(sys, 'frozen', False):
        script_dir = Path(sys.executable).parent
    else:
        script_dir = Path(__file__).parent

    setting_file = script_dir / "abnormal_setting.ini"
    if not setting_file.exists():
        print(f"錯誤: 找不到設定檔 {setting_file}")
        return None

    config = configparser.ConfigParser()
    config.read(setting_file, encoding='utf-8-sig')

    try:
        log_path    = config.get('logAnalyse', 'logPath').strip("'\"")
        output_path = config.get('logAnalyse', 'outputPath').strip("'\"")

        day_hr_start  = config.getint('logAnalyse', 'day_hr_start')
        day_min_start = config.getint('logAnalyse', 'day_min_start')
        day_hr_end    = config.getint('logAnalyse', 'day_hr_end')
        day_min_end   = config.getint('logAnalyse', 'day_min_end')

        night_hr_start  = config.getint('logAnalyse', 'night_hr_start')
        night_min_start = config.getint('logAnalyse', 'night_min_start')
        night_hr_end    = config.getint('logAnalyse', 'night_hr_end')
        night_min_end   = config.getint('logAnalyse', 'night_min_end')

        shift_config = {
            'day':   (day_hr_start, day_min_start, day_hr_end, day_min_end),
            'night': (night_hr_start, night_min_start, night_hr_end, night_min_end),
        }

        return {
            'log_path':     log_path,
            'output_path':  output_path,
            'shift_config': shift_config,
            'script_dir':   script_dir,
        }
    except (configparser.NoSectionError, configparser.NoOptionError) as e:
        print(f"錯誤: 設定檔格式錯誤 - {e}")
        return None


def calculate_shift_time_ranges(shift_config):
    """計算日班和夜班的時間範圍"""
    yesterday = datetime.now() - timedelta(days=1)
    today     = datetime.now()

    day_hr_start, day_min_start, day_hr_end, day_min_end = shift_config['day']
    day_start = yesterday.replace(hour=day_hr_start, minute=day_min_start, second=0, microsecond=0)
    day_end   = yesterday.replace(hour=day_hr_end,   minute=day_min_end,   second=0, microsecond=0)

    night_hr_start, night_min_start, night_hr_end, night_min_end = shift_config['night']
    night_start = yesterday.replace(hour=night_hr_start, minute=night_min_start, second=0, microsecond=0)
    night_end   = today.replace(hour=night_hr_end,       minute=night_min_end,   second=0, microsecond=0)

    return day_start, day_end, night_start, night_end


def get_yesterday_log_files(log_path, day_start, night_end):
    """取得前一天和今天的 log 檔案"""
    yesterday = datetime.now() - timedelta(days=1)
    today     = datetime.now()
    yesterday_prefix = yesterday.strftime("%Y%m%d")
    today_prefix     = today.strftime("%Y%m%d")

    log_folder = Path(log_path)
    if not log_folder.exists():
        print(f"錯誤: Log 路徑不存在: {log_path}")
        return [], yesterday_prefix

    log_files = []

    day_start_hour = day_start.hour
    for f in log_folder.glob(f"Log{yesterday_prefix}*.txt"):
        fname_body = f.stem[len("Log") + len(yesterday_prefix):]
        try:
            if int(fname_body) >= day_start_hour:
                log_files.append(f)
        except ValueError:
            log_files.append(f)

    night_end_hour = night_end.hour
    for f in log_folder.glob(f"Log{today_prefix}*.txt"):
        fname_body = f.stem[len("Log") + len(today_prefix):]
        try:
            if int(fname_body) <= night_end_hour:
                log_files.append(f)
        except ValueError:
            log_files.append(f)

    log_files.sort(key=lambda x: x.name)
    return log_files, yesterday_prefix


def move_logs_to_output(log_files, output_path, date_prefix):
    """將 log 檔案移動到輸出資料夾"""
    output_folder = Path(output_path) / f"Log{date_prefix}"
    output_folder.mkdir(parents=True, exist_ok=True)

    moved_files = []
    for log_file in log_files:
        dest_file = output_folder / log_file.name
        try:
            shutil.move(str(log_file), str(dest_file))
            moved_files.append(dest_file)
            print(f"  移動: {log_file.name}")
        except Exception as e:
            print(f"  警告: 無法移動 {log_file.name} - {e}")

    return output_folder, moved_files


def main():
    """主程式"""
    print("=" * 60)
    print("AGV 異常停留統計工具 (日班+夜班)")
    print("=" * 60)

    # 1. 讀取設定檔
    print("\n[1] 讀取設定檔...")
    settings = load_settings()
    if not settings:
        return 1

    log_path     = settings['log_path']
    output_path  = settings['output_path']
    shift_config = settings['shift_config']
    script_dir   = settings['script_dir']

    print(f"  Log 路徑: {log_path}")
    print(f"  輸出路徑: {output_path}")

    # 2. 載入點位設定（含全域異常閾值）
    print(f"\n[2] 載入點位設定...")
    point_settings         = load_point_settings(script_dir)
    excluded_points        = point_settings['excluded_points']
    elevator_points        = point_settings['elevator_points']
    elevator_threshold_sec = point_settings['elevator_threshold_sec']
    abnormal_threshold_min = point_settings['abnormal_threshold_min']
    abnormal_threshold_sec = abnormal_threshold_min * 60

    # 3. 計算時間範圍
    day_start, day_end, night_start, night_end = calculate_shift_time_ranges(shift_config)

    day_hours   = (day_end   - day_start).total_seconds()   / 3600
    night_hours = (night_end - night_start).total_seconds() / 3600

    print(f"\n[3] 統計時間範圍:")
    print(f"  日班: {day_start.strftime('%Y-%m-%d %H:%M')} ~ {day_end.strftime('%Y-%m-%d %H:%M')} ({day_hours:.1f}H)")
    print(f"  夜班: {night_start.strftime('%Y-%m-%d %H:%M')} ~ {night_end.strftime('%Y-%m-%d %H:%M')} ({night_hours:.1f}H)")

    # 4. 取得 log 檔案
    print(f"\n[4] 搜尋 Log 檔案...")
    log_files, date_prefix = get_yesterday_log_files(log_path, day_start, night_end)
    if not log_files:
        print(f"  找不到任何 Log 檔案")
        return 1
    print(f"  找到 {len(log_files)} 個 log 檔案")

    # 5. 移動檔案到輸出資料夾
    print(f"\n[5] 移動檔案到輸出資料夾...")
    output_folder, moved_files = move_logs_to_output(log_files, output_path, date_prefix)
    print(f"  輸出資料夾: {output_folder}")

    if not moved_files:
        print("  錯誤: 沒有成功移動任何檔案")
        return 1

    # 6. 解析 log（僅統計異常停留）
    print(f"\n[6] 解析 Log 檔案 (依結束時間歸屬班別)...")
    day_range   = (day_start,   day_end)
    night_range = (night_start, night_end)
    shift_results = parse_log_files_by_shift(
        moved_files, day_range, night_range, abnormal_threshold_sec,
        excluded_points, elevator_points, elevator_threshold_sec)

    # 7. 日班報告
    day_abnormal, day_seen_cars = shift_results['day']
    print_report(day_abnormal, len(moved_files), output_folder,
                 shift_label='日班', abnormal_threshold_min=abnormal_threshold_min)
    print(f"\n[7] 輸出日班 Excel 報告...")
    export_excel_abnormal(day_abnormal, day_seen_cars, output_folder,
                          shift_label='日班',
                          abnormal_threshold_min=abnormal_threshold_min,
                          date_prefix=date_prefix, shift_suffix='day')

    # 8. 夜班報告
    night_abnormal, night_seen_cars = shift_results['night']
    print_report(night_abnormal, len(moved_files), output_folder,
                 shift_label='夜班', abnormal_threshold_min=abnormal_threshold_min)
    print(f"\n[8] 輸出夜班 Excel 報告...")
    export_excel_abnormal(night_abnormal, night_seen_cars, output_folder,
                          shift_label='夜班',
                          abnormal_threshold_min=abnormal_threshold_min,
                          date_prefix=date_prefix, shift_suffix='night')

    print("\n" + "=" * 60)
    print("完成!")
    print("=" * 60)

    return 0


if __name__ == '__main__':
    sys.exit(main())
