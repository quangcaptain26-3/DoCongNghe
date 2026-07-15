# -*- coding: utf-8 -*-
"""Xuất báo cáo ra Excel (.xlsx) và CSV bằng openpyxl.

Điểm nhấn theo yêu cầu "báo cáo cho sếp":
  - Excel GẮN CÔNG THỨC SỐNG (SUM/ROUND/IF...) cho các ô tỷ lệ, cùng dòng tổng,
    để bấm vào ô là thấy cách tính và sửa đầu vào thì số tự đổi.
  - Có BIỂU ĐỒ (openpyxl.chart) ngay trong sheet Tổng quan.
  - CSV kèm cột đầu vào (giờ bất thường, số xe, mẫu số) + khối chú thích công thức,
    vì CSV không giữ được công thức sống một cách ổn định.

Toàn bộ chữ tiếng Việt CÓ DẤU.
"""

from __future__ import annotations

import csv
from datetime import date
from pathlib import Path
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .abnormal import DEFAULT_DENOM_HOURS, DayResult, ShiftResult
from .aggregate import (
    PeriodSummary,
    aggregate_cars,
    aggregate_points,
    daynight_split,
    monthly_summaries,
    overall_summary,
    saturday_summaries,
    severity_buckets,
    weekday_name,
    weekday_pattern,
    weekly_summaries,
)

# --- Style dùng chung ---------------------------------------------------------

_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_BOLD = Font(bold=True, size=11)
_NORMAL = Font(size=10)
_TITLE = Font(bold=True, size=15, color="1F4E79")
_SUBTITLE = Font(size=10, italic=True, color="666666")
_HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
_RATE_FILL = PatternFill("solid", fgColor="FCE4D6")
_KPI_FILL = PatternFill("solid", fgColor="E2EFDA")
_TOTAL_FILL = PatternFill("solid", fgColor="FFF2CC")
_THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

_PCT_FMT = '0.00"%"'
_HOUR_FMT = "0.00"

SHIFT_LABEL_VI = {"day": "Ca ngày", "night": "Ca đêm"}


def _style(cell, font=_NORMAL, alignment=_CENTER, border=_THIN, fill=None, number_format=None):
    cell.font = font
    cell.alignment = alignment
    cell.border = border
    if fill:
        cell.fill = fill
    if number_format:
        cell.number_format = number_format


def _set_widths(ws: Worksheet, widths: dict):
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width


def _write_header_row(ws: Worksheet, row: int, headers: List[str]) -> None:
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=i, value=h)
        _style(cell, font=_BOLD, fill=_HEADER_FILL)


# --- Sheet: Theo ngày (có công thức) ------------------------------------------

def _sheet_daily(ws: Worksheet, day_results: List[DayResult], denom_hours: float) -> int:
    """Ghi bảng theo ngày với công thức tỷ lệ. Trả về số dòng TỔNG."""
    ws.merge_cells("A1:M1")
    ws["A1"] = "Chất lượng AGV theo ngày (tỷ lệ = giờ bất thường / (số giờ mẫu số × số xe) × 100)"
    _style(ws["A1"], font=_TITLE, alignment=_LEFT)

    headers = [
        "Ngày", "Thứ", "Số xe", "Tổng lượt bất thường",
        "Giờ BT ca ngày", "Số xe ca ngày", "Giờ BT ca đêm", "Số xe ca đêm",
        "Số giờ mẫu số", "Tỷ lệ ca ngày (%)", "Tỷ lệ ca đêm (%)",
        "Tỷ lệ cả ngày (%)", "Số file log",
    ]
    _write_header_row(ws, 3, headers)

    row = 4
    first = row
    for d in sorted(day_results, key=lambda x: x.base_date):
        ws.cell(row=row, column=1, value=d.base_date.isoformat())
        ws.cell(row=row, column=2, value=weekday_name(d.base_date))
        ws.cell(row=row, column=3, value=d.car_count)
        ws.cell(row=row, column=4, value=d.abnormal_count)
        ws.cell(row=row, column=5, value=round(d.day.abnormal_hours, 4))
        ws.cell(row=row, column=6, value=d.day.car_count)
        ws.cell(row=row, column=7, value=round(d.night.abnormal_hours, 4))
        ws.cell(row=row, column=8, value=d.night.car_count)
        ws.cell(row=row, column=9, value=denom_hours)
        # Tỷ lệ = công thức sống
        ws.cell(row=row, column=10,
                value="=IF(F%d*I%d=0,0,ROUND(E%d/(I%d*F%d)*100,2))" % (row, row, row, row, row))
        ws.cell(row=row, column=11,
                value="=IF(H%d*I%d=0,0,ROUND(G%d/(I%d*H%d)*100,2))" % (row, row, row, row, row))
        ws.cell(row=row, column=12,
                value="=IF(C%d*I%d=0,0,ROUND((E%d+G%d)/(I%d*C%d)*100,2))"
                      % (row, row, row, row, row, row))
        ws.cell(row=row, column=13, value=d.log_file_count)

        for col in range(1, 14):
            fmt = _PCT_FMT if col in (10, 11, 12) else (_HOUR_FMT if col in (5, 7) else None)
            fill = _RATE_FILL if col == 12 else None
            _style(ws.cell(row=row, column=col), fill=fill, number_format=fmt)
        row += 1

    last = row - 1
    total = row
    ws.cell(row=total, column=1, value="TỔNG / CHUNG")
    ws.cell(row=total, column=3, value="=SUM(C%d:C%d)" % (first, last))
    ws.cell(row=total, column=4, value="=SUM(D%d:D%d)" % (first, last))
    ws.cell(row=total, column=5, value="=SUM(E%d:E%d)" % (first, last))
    ws.cell(row=total, column=7, value="=SUM(G%d:G%d)" % (first, last))
    ws.cell(row=total, column=9, value=denom_hours)
    ws.cell(row=total, column=12,
            value="=IF(C%d*I%d=0,0,ROUND((E%d+G%d)/(I%d*C%d)*100,2))"
                  % (total, total, total, total, total, total))
    for col in range(1, 14):
        fmt = _PCT_FMT if col == 12 else (_HOUR_FMT if col in (5, 7) else None)
        _style(ws.cell(row=total, column=col), font=_BOLD, fill=_TOTAL_FILL, number_format=fmt)

    _set_widths(ws, {
        "A": 12, "B": 10, "C": 8, "D": 16, "E": 14, "F": 13, "G": 14, "H": 13,
        "I": 13, "J": 15, "K": 15, "L": 15, "M": 10,
    })
    ws.freeze_panes = "A4"
    return total


# --- Sheet: theo kỳ (tuần / tháng / thứ 7) ------------------------------------

def _sheet_period(wb: Workbook, title: str, summaries: List[PeriodSummary],
                  denom_hours: float) -> None:
    ws = wb.create_sheet(title=title)
    ws.merge_cells("A1:J1")
    ws["A1"] = title + " (tỷ lệ = giờ bất thường / (số giờ mẫu số × số xe-ngày) × 100)"
    _style(ws["A1"], font=_TITLE, alignment=_LEFT)

    headers = [
        "Kỳ", "Từ ngày", "Đến ngày", "Số ngày", "Số xe (khác nhau)",
        "Số xe-ngày", "Tổng lượt bất thường", "Giờ bất thường",
        "Số giờ mẫu số", "Tỷ lệ bất thường (%)",
    ]
    _write_header_row(ws, 3, headers)

    row = 4
    for s in summaries:
        ws.cell(row=row, column=1, value=s.label)
        ws.cell(row=row, column=2, value=s.date_start.isoformat() if s.date_start else "-")
        ws.cell(row=row, column=3, value=s.date_end.isoformat() if s.date_end else "-")
        ws.cell(row=row, column=4, value=s.num_days)
        ws.cell(row=row, column=5, value=s.distinct_car_count)
        ws.cell(row=row, column=6, value=s.car_days)
        ws.cell(row=row, column=7, value=s.abnormal_count)
        ws.cell(row=row, column=8, value=round(s.abnormal_hours, 4))
        ws.cell(row=row, column=9, value=denom_hours)
        ws.cell(row=row, column=10,
                value="=IF(F%d*I%d=0,0,ROUND(H%d/(I%d*F%d)*100,2))" % (row, row, row, row, row))
        for col in range(1, 11):
            fmt = _PCT_FMT if col == 10 else (_HOUR_FMT if col == 8 else None)
            fill = _RATE_FILL if col == 10 else None
            _style(ws.cell(row=row, column=col), fill=fill, number_format=fmt)
        row += 1

    _set_widths(ws, {
        "A": 18, "B": 12, "C": 12, "D": 8, "E": 16, "F": 12, "G": 18,
        "H": 14, "I": 13, "J": 18,
    })


# --- Sheet: theo điểm ---------------------------------------------------------

def _sheet_points(wb: Workbook, day_results: List[DayResult]) -> Worksheet:
    ws = wb.create_sheet(title="Theo điểm")
    ws.merge_cells("A1:E1")
    ws["A1"] = "Xếp hạng ĐIỂM hay kẹt (điểm gây bất thường nhiều nhất)"
    _style(ws["A1"], font=_TITLE, alignment=_LEFT)

    _write_header_row(ws, 3, [
        "Điểm", "Số lượt bất thường", "Số xe liên quan",
        "Tổng thời gian (phút)", "Trung bình (phút)",
    ])
    row = 4
    for ps in aggregate_points(day_results):
        ws.cell(row=row, column=1, value="Điểm %s" % ps.point_id)
        ws.cell(row=row, column=2, value=ps.abnormal_count)
        ws.cell(row=row, column=3, value=ps.car_count)
        ws.cell(row=row, column=4, value=ps.abnormal_min)
        ws.cell(row=row, column=5, value="=IF(B%d=0,0,ROUND(D%d/B%d,2))" % (row, row, row))
        for col in range(1, 6):
            _style(ws.cell(row=row, column=col),
                   number_format=(_HOUR_FMT if col in (4, 5) else None))
        row += 1

    _set_widths(ws, {"A": 14, "B": 18, "C": 16, "D": 20, "E": 18})
    return ws


# --- Sheet: theo xe -----------------------------------------------------------

def _sheet_cars(wb: Workbook, day_results: List[DayResult]) -> Worksheet:
    ws = wb.create_sheet(title="Theo xe")
    ws.merge_cells("A1:D1")
    ws["A1"] = "Xếp hạng XE bất thường (toàn bộ dữ liệu đã nạp)"
    _style(ws["A1"], font=_TITLE, alignment=_LEFT)

    _write_header_row(ws, 3, [
        "Số xe", "Số lượt bất thường", "Tổng thời gian (phút)", "Trung bình (phút)",
    ])
    row = 4
    for cs in aggregate_cars(day_results):
        ws.cell(row=row, column=1, value=cs.car_id)
        ws.cell(row=row, column=2, value=cs.abnormal_count)
        ws.cell(row=row, column=3, value=cs.abnormal_min)
        ws.cell(row=row, column=4, value="=IF(B%d=0,0,ROUND(C%d/B%d,2))" % (row, row, row))
        for col in range(1, 5):
            _style(ws.cell(row=row, column=col),
                   number_format=(_HOUR_FMT if col in (3, 4) else None))
        row += 1

    _set_widths(ws, {"A": 12, "B": 18, "C": 22, "D": 18})
    return ws


# --- Sheet: chi tiết bất thường -----------------------------------------------

def _sheet_detail(wb: Workbook, day_results: List[DayResult]) -> None:
    ws = wb.create_sheet(title="Chi tiết bất thường")
    ws.merge_cells("A1:G1")
    ws["A1"] = "Chi tiết từng lần dừng bất thường"
    _style(ws["A1"], font=_TITLE, alignment=_LEFT)

    _write_header_row(ws, 3, [
        "Ngày", "Thứ", "Ca", "Số xe", "Điểm", "Giờ đến", "Thời gian dừng (phút)",
    ])
    row = 4
    first = row
    for d in sorted(day_results, key=lambda x: x.base_date):
        for shift in (d.day, d.night):
            by_car = shift.abnormal_by_car()
            for car_id in sorted(by_car.keys()):
                for rec in by_car[car_id]:
                    ws.cell(row=row, column=1, value=d.base_date.isoformat())
                    ws.cell(row=row, column=2, value=weekday_name(d.base_date))
                    ws.cell(row=row, column=3, value=SHIFT_LABEL_VI.get(shift.label, shift.label))
                    ws.cell(row=row, column=4, value=car_id)
                    ws.cell(row=row, column=5, value="Điểm %s" % rec.point_id)
                    ws.cell(row=row, column=6, value=rec.arrival_time.strftime("%H:%M:%S"))
                    ws.cell(row=row, column=7, value=rec.stay_min)
                    for col in range(1, 8):
                        _style(ws.cell(row=row, column=col),
                               number_format=(_HOUR_FMT if col == 7 else None))
                    row += 1

    last = row - 1
    if last >= first:
        ws.cell(row=row, column=6, value="Tổng thời gian (phút)")
        ws.cell(row=row, column=7, value="=SUM(G%d:G%d)" % (first, last))
        _style(ws.cell(row=row, column=6), font=_BOLD, fill=_TOTAL_FILL, alignment=_LEFT)
        _style(ws.cell(row=row, column=7), font=_BOLD, fill=_TOTAL_FILL, number_format=_HOUR_FMT)
    else:
        ws.cell(row=row, column=1, value="(Không có lần bất thường nào)")
        _style(ws.cell(row=row, column=1), font=_SUBTITLE, alignment=_LEFT, border=Border())

    _set_widths(ws, {"A": 12, "B": 10, "C": 10, "D": 8, "E": 12, "F": 12, "G": 20})
    ws.freeze_panes = "A4"


# --- Sheet: Thứ 7 chi tiết ----------------------------------------------------

def _sheet_saturday_detail(wb: Workbook, day_results: List[DayResult]) -> None:
    saturdays = [d for d in day_results if d.is_saturday]
    if not saturdays:
        return
    ws = wb.create_sheet(title="Thứ 7 chi tiết")
    ws.merge_cells("A1:E1")
    ws["A1"] = "Chi tiết bất thường ngày Thứ Bảy (AGV chạy khi nghỉ)"
    _style(ws["A1"], font=_TITLE, alignment=_LEFT)

    _write_header_row(ws, 3, ["Ngày", "Ca", "Số xe", "Điểm / Giờ", "Thời gian dừng (phút)"])
    row = 4
    for d in sorted(saturdays, key=lambda x: x.base_date):
        for shift in (d.day, d.night):
            by_car = shift.abnormal_by_car()
            for car_id in sorted(by_car.keys()):
                for rec in by_car[car_id]:
                    ws.cell(row=row, column=1, value=d.base_date.isoformat())
                    ws.cell(row=row, column=2, value=SHIFT_LABEL_VI.get(shift.label, shift.label))
                    ws.cell(row=row, column=3, value=car_id)
                    ws.cell(row=row, column=4,
                            value="Điểm %s / %s" % (rec.point_id, rec.arrival_time.strftime("%H:%M:%S")))
                    ws.cell(row=row, column=5, value=rec.stay_min)
                    for col in range(1, 6):
                        _style(ws.cell(row=row, column=col),
                               number_format=(_HOUR_FMT if col == 5 else None))
                    row += 1
    _set_widths(ws, {"A": 14, "B": 10, "C": 8, "D": 26, "E": 20})


# --- Sheet: Tổng quan (dashboard + biểu đồ) -----------------------------------

def _sheet_dashboard(ws: Worksheet, day_results: List[DayResult], denom_hours: float,
                     daily_total_row: int, point_last_row: int) -> None:
    summ = overall_summary(day_results, denom_hours)

    ws.merge_cells("A1:F1")
    ws["A1"] = "BÁO CÁO CHẤT LƯỢNG HOẠT ĐỘNG AGV"
    _style(ws["A1"], font=Font(bold=True, size=17, color="1F4E79"), alignment=_LEFT, border=Border())

    ws.merge_cells("A2:F2")
    date_range = "-"
    if summ.worst_day and day_results:
        d0 = min(d.base_date for d in day_results)
        d1 = max(d.base_date for d in day_results)
        date_range = "%s -> %s" % (d0.isoformat(), d1.isoformat())
    ws["A2"] = "Khoảng dữ liệu: %s   |   Số giờ mẫu số: %g giờ/ngày" % (date_range, denom_hours)
    _style(ws["A2"], font=_SUBTITLE, alignment=_LEFT, border=Border())

    # Khối KPI (2 cột nhãn:giá trị). Các ô tỷ lệ/tổng dùng công thức sống.
    dt = daily_total_row
    kpis = [
        ("Số ngày phân tích", summ.num_days, None),
        ("Tổng số xe (khác nhau)", summ.distinct_cars, None),
        ("Tổng lượt bất thường", "='Theo ngày'!D%d" % dt, None),
        ("Tổng giờ bất thường", "='Theo ngày'!E%d+'Theo ngày'!G%d" % (dt, dt), _HOUR_FMT),
        ("Tỷ lệ bất thường CHUNG (%)", "='Theo ngày'!L%d" % dt, _PCT_FMT),
        ("Trung bình lượt/ngày", "=IF(%d=0,0,ROUND('Theo ngày'!D%d/%d,2))"
            % (summ.num_days, dt, summ.num_days) if summ.num_days else 0, _HOUR_FMT),
    ]
    r = 4
    for i, (label, value, fmt) in enumerate(kpis):
        base_col = 1 if i % 2 == 0 else 4
        if i % 2 == 0 and i > 0:
            r += 1
        lc = ws.cell(row=r, column=base_col, value=label)
        vc = ws.cell(row=r, column=base_col + 1, value=value)
        _style(lc, font=_BOLD, fill=_KPI_FILL, alignment=_LEFT)
        _style(vc, font=Font(bold=True, size=12, color="C00000"), fill=_KPI_FILL,
               number_format=fmt)

    # Điểm nhấn (worst)
    r += 2
    highlights = [
        ("Ngày tệ nhất",
         ("%s (%s) - %.2f%%" % (summ.worst_day.base_date.isoformat(),
                                weekday_name(summ.worst_day.base_date),
                                summ.worst_day.abnormal_rate(denom_hours))
          ) if summ.worst_day else "-"),
        ("Ngày tốt nhất",
         ("%s (%s) - %.2f%%" % (summ.best_day.base_date.isoformat(),
                                weekday_name(summ.best_day.base_date),
                                summ.best_day.abnormal_rate(denom_hours))
          ) if summ.best_day else "-"),
        ("Xe cần chú ý nhất",
         ("Xe %s - %d lượt (%.1f phút)" % (summ.top_car.car_id, summ.top_car.abnormal_count,
                                           summ.top_car.abnormal_min)) if summ.top_car else "-"),
        ("Điểm hay kẹt nhất",
         ("Điểm %s - %d lượt (%d xe)" % (summ.top_point.point_id, summ.top_point.abnormal_count,
                                         summ.top_point.car_count)) if summ.top_point else "-"),
    ]
    for label, value in highlights:
        lc = ws.cell(row=r, column=1, value=label)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        vc = ws.cell(row=r, column=2, value=value)
        _style(lc, font=_BOLD, alignment=_LEFT)
        _style(vc, alignment=_LEFT)
        r += 1

    _set_widths(ws, {"A": 26, "B": 20, "C": 14, "D": 24, "E": 16, "F": 10})

    # --- Biểu đồ: tỷ lệ bất thường theo ngày ---
    ws_daily = ws.parent["Theo ngày"]
    daily_last_data = daily_total_row - 1
    if daily_last_data >= 4:
        chart = LineChart()
        chart.title = "Tỷ lệ bất thường theo ngày (%)"
        chart.height = 8
        chart.width = 18
        chart.y_axis.title = "%"
        data = Reference(ws_daily, min_col=12, min_row=3, max_row=daily_last_data)
        cats = Reference(ws_daily, min_col=1, min_row=4, max_row=daily_last_data)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "A%d" % (r + 1))

    # --- Biểu đồ: top điểm hay kẹt ---
    if point_last_row >= 4:
        ws_points = ws.parent["Theo điểm"]
        top_last = min(point_last_row, 13)   # tối đa 10 điểm đầu
        bar = BarChart()
        bar.type = "bar"
        bar.title = "Top điểm hay kẹt (số lượt bất thường)"
        bar.height = 8
        bar.width = 18
        data = Reference(ws_points, min_col=2, min_row=3, max_row=top_last)
        cats = Reference(ws_points, min_col=1, min_row=4, max_row=top_last)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        ws.add_chart(bar, "A%d" % (r + 18))


# --- Điểm vào chính -----------------------------------------------------------

def export_full_report(day_results: List[DayResult], out_path: Path,
                       threshold_min: int = 12,
                       denom_hours: float = DEFAULT_DENOM_HOURS) -> Path:
    """Xuất workbook tổng hợp nhiều sheet (công thức sống + biểu đồ)."""
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws_dash = wb.active
    ws_dash.title = "Tổng quan"

    ws_daily = wb.create_sheet(title="Theo ngày")
    daily_total_row = _sheet_daily(ws_daily, day_results, denom_hours)

    _sheet_period(wb, "Theo tuần", weekly_summaries(day_results, denom_hours), denom_hours)
    _sheet_period(wb, "Theo tháng", monthly_summaries(day_results, denom_hours), denom_hours)

    sat = saturday_summaries(day_results, denom_hours)
    if sat:
        _sheet_period(wb, "Thứ 7", sat, denom_hours)
        _sheet_saturday_detail(wb, day_results)

    ws_points = _sheet_points(wb, day_results)
    point_last_row = ws_points.max_row
    _sheet_cars(wb, day_results)
    _sheet_detail(wb, day_results)

    _sheet_dashboard(ws_dash, day_results, denom_hours, daily_total_row, point_last_row)

    wb.save(out_path)
    return out_path


# --- Định dạng gốc (parity với ptich_agv, tỷ lệ là công thức) -----------------

def write_original_shift_sheet(ws: Worksheet, shift: ShiftResult, shift_label_vi: str,
                               threshold_min: int, denom_hours: float = DEFAULT_DENOM_HOURS) -> None:
    """Ghi một ca theo đúng bố cục xlsx bản gốc: B=xe, C=số lần, D=điểm/giờ, E=phút.

    Ô tổng tỷ lệ (C3) dùng CÔNG THỨC thay vì chuỗi tĩnh để đối chiếu được.
    """
    ws.merge_cells("B1:E1")
    c = ws["B1"]
    c.value = "AGV xe - Thống kê thời gian bất thường (%s)" % shift_label_vi
    _style(c, font=Font(bold=True, size=14))

    ws["B3"] = "Tổng tỷ lệ bất thường (O)"
    _style(ws["B3"], font=_BOLD)
    ws.merge_cells("D3:E3")
    ws["D3"] = "*Tổng (E) của mọi xe / (số giờ mẫu số × tổng số xe)"
    _style(ws["D3"], font=Font(size=9, italic=True))

    ws["C5"] = "Số lần dừng bất thường"
    ws["D5"] = "Điểm bất thường"
    ws["E5"] = "Thời gian lưu bất thường"
    for col in ("C", "D", "E"):
        _style(ws["%s5" % col], font=_BOLD, fill=_HEADER_FILL)

    ws["B7"] = "Số xe \\ Định nghĩa"
    ws["C7"] = "Dừng quá %d phút (trừ sạc)" % threshold_min
    ws["D7"] = "Ghi chú điểm bất thường (1 điểm/1 giờ/1 dòng)"
    ws["E7"] = "Thời gian lưu tại điểm bất thường (phút)"
    for col in ("B", "C", "D", "E"):
        _style(ws["%s7" % col], font=_BOLD, fill=_HEADER_FILL)

    by_car = shift.abnormal_by_car()
    all_cars = shift.all_cars
    row = 8
    first = row
    for car_id in all_cars:
        records = by_car.get(car_id, [])
        if records:
            for j, rec in enumerate(records):
                if j == 0:
                    ws.cell(row=row, column=2, value=car_id)
                    ws.cell(row=row, column=3, value=len(records))
                ws.cell(row=row, column=4,
                        value="Điểm%s / %s" % (rec.point_id, rec.arrival_time.strftime("%H:%M:%S")))
                ws.cell(row=row, column=5, value=rec.stay_min)
                for col in range(2, 6):
                    _style(ws.cell(row=row, column=col),
                           number_format=(_HOUR_FMT if col == 5 else None))
                row += 1
        else:
            ws.cell(row=row, column=2, value=car_id)
            ws.cell(row=row, column=3, value=0)
            ws.cell(row=row, column=4, value="-")
            ws.cell(row=row, column=5, value="-")
            for col in range(2, 6):
                _style(ws.cell(row=row, column=col))
            row += 1
    last = row - 1

    car_count = len(all_cars)
    if car_count > 0 and last >= first:
        ws["C3"] = ("=ROUND(SUM(E%d:E%d)/60/(%g*%d)*100,2)"
                    % (first, last, denom_hours, car_count))
    else:
        ws["C3"] = 0
    _style(ws["C3"], font=Font(bold=True, size=11, color="FF0000"), fill=_RATE_FILL,
           number_format=_PCT_FMT)

    note_row = row + 2
    ws.cell(row=note_row, column=2,
            value="Ngưỡng bất thường: %d phút (trừ sạc)" % threshold_min)
    _style(ws.cell(row=note_row, column=2),
           font=Font(size=9, italic=True, color="666666"),
           alignment=_LEFT, border=Border())

    _set_widths(ws, {"A": 2, "B": 12, "C": 20, "D": 30, "E": 22})


def export_original_shift(shift: ShiftResult, base_date: date, out_dir: Path,
                          threshold_min: int, shift_suffix: str,
                          denom_hours: float = DEFAULT_DENOM_HOURS) -> Path:
    """Tái tạo file Log{date}_{day|night}.xlsx theo định dạng gốc."""
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_file = out_dir / ("Log%s_%s.xlsx" % (base_date.strftime("%Y%m%d"), shift_suffix))
    wb = Workbook()
    ws = wb.active
    ws.title = "daily"
    write_original_shift_sheet(ws, shift, SHIFT_LABEL_VI.get(shift.label, shift.label),
                               threshold_min, denom_hours)
    wb.save(out_file)
    return out_file


# --- CSV (giá trị + đầu vào + chú thích công thức) ----------------------------

def export_summary_csv(day_results: List[DayResult], out_path: Path,
                       denom_hours: float = DEFAULT_DENOM_HOURS) -> Path:
    """Xuất CSV tổng quan.

    Vì CSV không giữ được công thức sống ổn định (Excel tiếng Việt dùng ';'), file này
    kèm ĐẦY ĐỦ CỘT ĐẦU VÀO để tự tính lại + khối chú thích công thức bằng chữ.
    """
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    with open(out_path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)

        w.writerow(["BÁO CÁO CHẤT LƯỢNG HOẠT ĐỘNG AGV (tổng quan theo ngày)"])
        w.writerow([])

        # Khối chú thích công thức
        w.writerow(["CÔNG THỨC TÍNH"])
        w.writerow(["Tỷ lệ ca ngày (%)", "= Giờ BT ca ngày / (Số giờ mẫu số × Số xe ca ngày) × 100"])
        w.writerow(["Tỷ lệ ca đêm (%)", "= Giờ BT ca đêm / (Số giờ mẫu số × Số xe ca đêm) × 100"])
        w.writerow(["Tỷ lệ cả ngày (%)",
                    "= (Giờ BT ca ngày + Giờ BT ca đêm) / (Số giờ mẫu số × Số xe) × 100"])
        w.writerow(["Số giờ mẫu số hiện dùng", denom_hours])
        w.writerow([])

        # Bảng dữ liệu (kèm cột đầu vào để kiểm chứng)
        w.writerow([
            "Ngày", "Thứ", "Số xe", "Tổng lượt bất thường",
            "Giờ BT ca ngày", "Số xe ca ngày", "Giờ BT ca đêm", "Số xe ca đêm",
            "Số giờ mẫu số",
            "Tỷ lệ ca ngày (%)", "Tỷ lệ ca đêm (%)", "Tỷ lệ cả ngày (%)",
            "Số file log",
        ])
        for d in sorted(day_results, key=lambda x: x.base_date):
            w.writerow([
                d.base_date.isoformat(),
                weekday_name(d.base_date),
                d.car_count,
                d.abnormal_count,
                round(d.day.abnormal_hours, 4),
                d.day.car_count,
                round(d.night.abnormal_hours, 4),
                d.night.car_count,
                denom_hours,
                d.day.abnormal_rate(denom_hours),
                d.night.abnormal_rate(denom_hours),
                d.abnormal_rate(denom_hours),
                d.log_file_count,
            ])

        # Xếp hạng điểm
        w.writerow([])
        w.writerow(["XẾP HẠNG ĐIỂM HAY KẸT"])
        w.writerow(["Điểm", "Số lượt bất thường", "Số xe liên quan",
                    "Tổng thời gian (phút)", "Trung bình (phút)"])
        for ps in aggregate_points(day_results):
            w.writerow(["Điểm %s" % ps.point_id, ps.abnormal_count, ps.car_count,
                        ps.abnormal_min, ps.avg_min])

        # Xếp hạng xe
        w.writerow([])
        w.writerow(["XẾP HẠNG XE BẤT THƯỜNG"])
        w.writerow(["Số xe", "Số lượt bất thường", "Tổng thời gian (phút)", "Trung bình (phút)"])
        for cs in aggregate_cars(day_results):
            w.writerow([cs.car_id, cs.abnormal_count, cs.abnormal_min, cs.avg_min])

        # So sánh ca ngày / đêm
        w.writerow([])
        w.writerow(["SO SÁNH CA NGÀY / CA ĐÊM"])
        w.writerow(["Ca", "Số lượt bất thường", "Tổng thời gian (phút)",
                    "Số xe-ngày", "Tỷ lệ bất thường (%)"])
        for agg in daynight_split(day_results, denom_hours):
            w.writerow([agg.label, agg.abnormal_count, agg.abnormal_min,
                        agg.car_days, agg.abnormal_rate])

        # Mẫu theo thứ
        w.writerow([])
        w.writerow(["MẪU THEO THỨ TRONG TUẦN"])
        w.writerow(["Thứ", "Số ngày", "Số lượt bất thường", "Số xe-ngày", "Tỷ lệ bất thường (%)"])
        for ws_stat in weekday_pattern(day_results, denom_hours):
            w.writerow([ws_stat.name, ws_stat.num_days, ws_stat.abnormal_count,
                        ws_stat.car_days, ws_stat.abnormal_rate])

        # Phân nhóm mức độ nặng
        w.writerow([])
        w.writerow(["PHÂN NHÓM MỨC ĐỘ NẶNG (theo thời gian dừng)"])
        w.writerow(["Nhóm", "Số lượt", "Tổng thời gian (phút)"])
        for b in severity_buckets(day_results):
            w.writerow([b.label, b.count, b.abnormal_min])

    return out_path
